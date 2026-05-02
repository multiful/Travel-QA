"""322개 장소 운영시간 baseline + 카테고리 표준값 세팅.

산출물:
    1. gun/data/place_hours_322.xlsx  — 322개 모두에 운영시간 매핑
    2. gun/data/category_hours.xlsx   — 카테고리·sub_category별 표준 운영시간
    3. gun/src/data/hours_db.py       — 코드 룩업 모듈 (dwell_db.py 패턴)

설계:
    1. 카테고리힌트 (관광지/문화시설/식당/쇼핑/...) 1차 분류
    2. 여행지명 키워드 매칭으로 sub_category 도출
    3. sub_category → (open, break_start, break_end, close, off_days) 매핑
    4. 매칭 안 되면 카테고리 default 폴백
    5. confidence 라벨 (high / medium / low) 부여 — 후속 검색 우선순위
"""
from __future__ import annotations

import re
from pathlib import Path
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
INPUT_XLSX  = ROOT / "gun/data/recommendations_input.xlsx"
OUT_PLACE   = ROOT / "gun/data/place_hours_322.xlsx"
OUT_STD     = ROOT / "gun/data/category_hours.xlsx"
OUT_PY      = ROOT / "gun/src/data/hours_db.py"

# ─────────────────────────────────────────────────────────────────
# Sub-category 정의 — 한국 일반 운영시간 (2025 기준 표준)
#   open / break_start / break_end / close / off_days
#   off_days: 0=월, 1=화, ..., 6=일,  None = 휴무 없음
# ─────────────────────────────────────────────────────────────────
SUB_CATEGORIES: dict[str, dict] = {
    # ── 자연 (24시간 운영) ──────────────────────────────────────
    "nature_outdoor": {
        "label": "자연 · 산·해변·공원",
        "open": "00:00", "break_start": None, "break_end": None,
        "close": "23:59", "off_days": [],
        "note": "야외 자연지, 24시간 자유 출입 가정",
    },
    # ── 궁궐 (월·화 중 하나 휴무) ───────────────────────────────
    "palace": {
        "label": "궁궐 · 종묘",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [1],   # 화요일 (창덕궁 등)
        "note": "월요일 휴무 = 경복궁, 화요일 휴무 = 창덕궁/종묘 — 검색 보정 권장",
    },
    "tomb_historic": {
        "label": "왕릉 · 유적지 · 성곽",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [0],   # 월요일
        "note": "조선 왕릉 등 사적지는 월요일 휴무",
    },
    "temple": {
        "label": "사찰 · 사찰림",
        "open": "04:00", "break_start": None, "break_end": None,
        "close": "19:00", "off_days": [],
        "note": "사찰은 새벽 예불부터 일몰까지",
    },
    "traditional_village": {
        "label": "한옥마을 · 민속촌",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "마을 자체는 24시간이나 시설은 09–18시",
    },

    # ── 거리·광장·시장 (24시간 또는 새벽~밤) ───────────────────
    "street_plaza": {
        "label": "거리 · 광장 · 골목",
        "open": "00:00", "break_start": None, "break_end": None,
        "close": "23:59", "off_days": [],
        "note": "공공 공간, 24시간 출입",
    },
    "market_traditional": {
        "label": "전통시장",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "지역 전통시장 일반 시간 (개별 가게 편차 큼)",
    },
    "market_seafood": {
        "label": "수산시장",
        "open": "04:00", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "새벽 경매 04–06시, 횟집은 22시까지",
    },
    "night_market": {
        "label": "야시장",
        "open": "18:00", "break_start": None, "break_end": None,
        "close": "23:00", "off_days": [],
        "note": "야시장은 저녁부터",
    },

    # ── 전망대·타워 ────────────────────────────────────────────
    "tower_observatory": {
        "label": "타워 · 전망대",
        "open": "10:00", "break_start": None, "break_end": None,
        "close": "22:00", "off_days": [],
        "note": "주요 타워 일반 운영시간",
    },

    # ── 박물관·미술관·문화시설 (월요일 휴무) ───────────────────
    "museum": {
        "label": "박물관 · 미술관 · 기념관",
        "open": "10:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [0],   # 월요일
        "note": "공·사립 대부분 월요일 휴무",
    },
    "library_culture_center": {
        "label": "도서관 · 문화회관",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [0],
        "note": "공공 도서관 주요 시간",
    },
    "theme_park": {
        "label": "테마파크 · 놀이공원",
        "open": "10:00", "break_start": None, "break_end": None,
        "close": "22:00", "off_days": [],
        "note": "에버랜드/롯데월드 등",
    },
    "aquarium_zoo": {
        "label": "수족관 · 동물원",
        "open": "09:30", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "코엑스아쿠아리움/씨라이프 등",
    },

    # ── 식당 ───────────────────────────────────────────────────
    "korean_restaurant": {
        "label": "한식당",
        "open": "11:00", "break_start": "15:00", "break_end": "17:00",
        "close": "21:30", "off_days": [],
        "note": "한정식·국밥·삼계탕 일반 (브레이크 있음)",
    },
    "bbq_restaurant": {
        "label": "고기집 · 바비큐",
        "open": "11:30", "break_start": None, "break_end": None,
        "close": "23:00", "off_days": [],
        "note": "갈비·삼겹살, 브레이크 없는 곳 많음",
    },
    "noodle_restaurant": {
        "label": "면 · 분식",
        "open": "10:30", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "냉면·국수·분식",
    },
    "general_restaurant": {
        "label": "일반 식당",
        "open": "11:00", "break_start": "15:00", "break_end": "17:00",
        "close": "21:00", "off_days": [],
        "note": "기본 식당 표준값",
    },

    # ── 카페 ───────────────────────────────────────────────────
    "cafe": {
        "label": "카페 · 디저트",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "22:00", "off_days": [],
        "note": "프랜차이즈·로컬 카페 일반",
    },
    "bakery": {
        "label": "베이커리",
        "open": "08:00", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "빵집 일반",
    },

    # ── 쇼핑 ───────────────────────────────────────────────────
    "department_store": {
        "label": "백화점 · 쇼핑몰",
        "open": "10:30", "break_start": None, "break_end": None,
        "close": "20:00", "off_days": [],
        "note": "매장은 21시까지, 식당가는 22시까지인 곳 많음",
    },
    "outlet": {
        "label": "아울렛",
        "open": "10:30", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "프리미엄 아울렛 표준",
    },

    # ── 레포츠·체험 ────────────────────────────────────────────
    "leisure_sport": {
        "label": "레포츠 · 체험",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "패러글라이딩·서핑·체험농장 등",
    },

    # ── 숙박 (depot용) ────────────────────────────────────────
    "lodging": {
        "label": "숙박",
        "open": "00:00", "break_start": None, "break_end": None,
        "close": "23:59", "off_days": [],
        "note": "체크인 15:00 / 체크아웃 11:00 별도 룰",
    },

    # ── 종교시설 (성당·교회) ─────────────────────────────────
    "religious_site": {
        "label": "성당 · 교회",
        "open": "06:00", "break_start": None, "break_end": None,
        "close": "21:00", "off_days": [],
        "note": "미사·예배 시간 외 자유 출입, 야간 폐문 곳 있음",
    },

    # ── 명승지 (자연·문화 복합) ─────────────────────────────
    "scenic_spot": {
        "label": "명승지 · 절경",
        "open": "00:00", "break_start": None, "break_end": None,
        "close": "23:59", "off_days": [],
        "note": "외돌개·주상절리·정동진 등 야외 명승지",
    },

    # ── 동굴 (입장료 있는 관광 동굴) ────────────────────────
    "cave": {
        "label": "동굴 (관광)",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "만장굴·고씨굴·천지연 등 입장권 동굴",
    },

    # ── 케이블카·관광 교통 ──────────────────────────────────
    "cable_car": {
        "label": "케이블카 · 모노레일",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "기상 악화 시 운휴, 주말은 21시까지인 곳도",
    },

    # ── 목장·체험농장 ────────────────────────────────────────
    "ranch_farm": {
        "label": "목장 · 체험농장",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "대관령양떼목장 등, 동절기 17시 종료",
    },

    # ── 기타·미분류 ────────────────────────────────────────────
    "default_attraction": {
        "label": "기타 관광지 (미분류)",
        "open": "09:00", "break_start": None, "break_end": None,
        "close": "18:00", "off_days": [],
        "note": "베이스라인 — 검색 보정 권장",
    },
}


# ─────────────────────────────────────────────────────────────────
# 키워드 → sub_category 매칭 룰 (위에서 아래 순서로 평가)
# ─────────────────────────────────────────────────────────────────
KEYWORD_RULES: list[tuple[str, str]] = [
    # ── 1. 사찰 (화이트리스트) — 절벽인 "낙화암" 등 오탐 방지 ───
    (r"봉은사|해동용궁사|불국사|석굴암|낙산사|월정사|향일암|선암사|구인사|마곡사|"
     r"전등사|진관사|용궁사|부석사|보리암|쌍계사|선운사|용문사|해인사|통도사|"
     r"범어사|송광사|법주사|미황사|화엄사|쌍봉사|운주사|금산사|내장사|관음사|"
     r"약천사|약사사|봉정사|봉선사|적천사|자장사|봉선사|동학사|갑사|마이산탑사|"
     r"청룡사|봉선사|관룡사|사불산|불교문화관", "temple"),

    # ── 2. 동굴·관광 터널 ────────────────────────────────────
    (r"만장굴|고씨굴|천지연.*동굴|화암동굴|단양고수동굴|고수동굴|환선굴|여천굴|"
     r"동굴$|관광동굴|대금굴|해저터널|관광터널", "cave"),

    # ── 3. 케이블카·모노레일·짚라인 ───────────────────────────
    (r"케이블카|모노레일|곤돌라|로프웨이|루지$|짚라인|집라인|레일바이크", "cable_car"),

    # ── 4. 목장·체험농장 ──────────────────────────────────────
    (r"목장|양떼목장|허브농장|농원$|체험농장|승마장", "ranch_farm"),

    # ── 5. 자연 (산·바다·공원·숲·섬·저수지·정원) ────────────
    (r"산$|^[가-힣]+산\s|산림욕|봉우리|봉$|폭포|계곡|숲$|호수|호$|강$|바다|"
     r"해수욕장|해변$|해안|해변로|해돋이|일출|일몰|만$|곶$|섬$|비치$|"
     r"^.+도$|^오동도|^외도|^청산도|^보길도|^외연도|^안면도|^거문도|^월미도|^남이섬|"
     r"공원|숲길|올레|둘레길|등산로|언덕|고개|평원|벌판|초원|꽃밭|들꽃|"
     r"수목원|식물원|지질공원|자연휴양림|휴양림|자연공원|국립공원|도립공원|"
     r"가든|편백|편백숲|핀크스|메타세콰이어|"
     r"소쇄원|죽녹원|관방제림|보성녹차밭|녹차밭|식물원|"
     r"저수지$|^.+호$|의림지|화진포|두물머리|세미원|태화강국가정원|"
     r"방조림|어부림|^.+림$|"
     r"불광천|청계천|^.+천$|"
     r"동촌유원지|유원지$"
     , "nature_outdoor"),

    # ── 6. 명승지·절경 (자연 명소, 첨성대·오목대 등 유적은 별도) ──
    (r"낙화암|외돌개|주상절리|정방폭포|천지연|용두암|섭지코지|성산일출봉|"
     r"비자림|사려니숲|에코랜드|카멜리아힐|선유도|"
     r"영금정|공지천|경포대|정동진|월정리|함덕|애월|"
     r"순천만습지|순천만|우포늪|화담숲|아침고요|남이섬|쁘띠프랑스|"
     r"대왕암|간절곶|호미곶|가천다랭이|"
     r"용궁리|섭지|올레길|둘레길|"
     r"태종대|갓바위|신선대|촛대바위|하조대|화진포|정선아우라지|탄금대|"
     r"통일전망대|"
     r"묵호항|구룡포항|항구$"
     , "scenic_spot"),

    # ── 7. 궁궐 / 종묘 ────────────────────────────────────────
    (r"경복궁|창덕궁|창경궁|덕수궁|경희궁|운현궁|종묘", "palace"),

    # ── 8. 왕릉·유적·성·서원·향교·정자·누각 ──────────────────
    (r"왕릉|^선릉|^정릉|^태릉|^홍릉|^동구릉|능원|능$|고분|고도$|"
     r"유적|유적지|읍성|산성|토성|주성|남한산성|북한산성|고려궁지|문화유산|"
     r"서원$|향교$|선비촌|"
     r"경기전|오죽헌|동궁과월지|대릉원|노서리|쪽샘|첨성대|오목대|월영교|"
     r"수원화성|행궁$|행궁가|"
     r"^.+문$|숭례문|흥인지문|독립문|광화문|풍남문|팔달문|장안문|동대문$|남대문$|"
     r"^.+루$|광한루|촉석루|죽서루|만대루|영남루|"
     r"^.+사지$|정림사지", "tomb_historic"),

    # ── 9. 한옥마을·민속촌·전통마을 ───────────────────────────
    (r"한옥마을|민속촌|전통마을|민속마을|문화마을|북촌$|서촌$|감천문화마을|"
     r"양동마을|하회마을|외암마을|왕곡마을|전주한옥|전주한옥마을", "traditional_village"),

    # ── 10. 야시장 ────────────────────────────────────────────
    (r"야시장|밤시장", "night_market"),

    # ── 11. 수산시장·어시장 ──────────────────────────────────
    (r"수산시장|어시장|자갈치|소래포구|구룡포|주문진|속초중앙시장.*어시장", "market_seafood"),

    # ── 12. 전통시장 ──────────────────────────────────────────
    (r"시장$|전통시장|광장시장|남대문시장|동대문시장|통인시장|망원시장|"
     r"국제시장|부평시장|중앙시장|BIFF광장|깡통시장|남포동$", "market_traditional"),

    # ── 13. 백화점·복합몰·면세점 ─────────────────────────────
    (r"백화점|면세점|쇼핑몰|몰$|복합쇼핑|스타필드|이마트|코엑스|롯데월드몰|"
     r"가로수길.*몰|타임스퀘어|디큐브|디큐브시티|롯데몰|현대시티몰", "department_store"),

    # ── 14. 아울렛 ────────────────────────────────────────────
    (r"아울렛|프리미엄아울렛|시즌아울렛", "outlet"),

    # ── 15. 핫플레이스 거리·동네 (street/plaza 보강) ─────────
    (r"광화문|시청$|청계천|성수동|연남동|이태원|망원동|북촌$|서촌$|을지로$|"
     r"가로수길|경리단길|샤로수길|망원|연희동|"
     r"인사동|삼청동|익선동|서순라길|남대문로|"
     r"홍대|홍익대|건대$|건대거리|신촌$|"
     r"해방촌|용산공예단지|"
     r"부산.*해리단길|해리단길|"
     r"제주.*탑동|탑동광장", "street_plaza"),

    # ── 16. 일반 거리·광장·골목 ──────────────────────────────
    (r"거리$|^.+거리\s|^.+로$|^.+길$|골목$|광장$|마을$|동네|로데오|단지$|"
     r"상점가|쇼핑거리|먹자골목|예술의거리|벽화마을", "street_plaza"),

    # ── 17. 타워·전망대 ──────────────────────────────────────
    (r"타워|전망대|뷰포인트|스카이|루프탑|전망$|^N서울|남산서울타워", "tower_observatory"),

    # ── 18. 박물관·미술관·뮤지엄·기념관 ──────────────────────
    (r"박물관|미술관|뮤지엄|museum|기념관|아트센터|아트홀|문학관|역사관|"
     r"체험관|과학관|전시관|갤러리|예술관|디자인플라자|DDP|"
     r"테디베어|티뮤지엄|오설록|스페이스워크|아트밸리|포천아트밸리", "museum"),

    # ── 19. 도서관·문화회관 ──────────────────────────────────
    (r"도서관|문화회관|문화원|아트빌리지|예술의전당", "library_culture_center"),

    # ── 20. 테마파크 ─────────────────────────────────────────
    (r"에버랜드|롯데월드$|레전드$|월드$|테마파크|놀이공원|랜드$|어드벤처|"
     r"한국민속촌|드림랜드|월미테마파크", "theme_park"),

    # ── 21. 수족관·동물원·아쿠아 ─────────────────────────────
    (r"아쿠아리움|수족관|아쿠아|씨라이프|동물원|사파리|아쿠아필드", "aquarium_zoo"),

    # ── 22. 종교시설 (성당·교회) ─────────────────────────────
    (r"성당$|교회$|채플|수도원|^명동성당|약현성당|중림동성당", "religious_site"),

    # ── 23. 식당 ─────────────────────────────────────────────
    (r"갈비|삼겹살|돼지|소고기|등심|안심|불고기|고기집|숯불|화로구이", "bbq_restaurant"),
    (r"국밥|순대|냉면|국수|면옥|분식|떡볶이|김밥|만두|짜장면|짬뽕", "noodle_restaurant"),
    (r"한정식|한식|삼계탕|곰탕|설렁탕|족발|보쌈|닭갈비|찜닭|토속촌|"
     r"광장시장.*빈대떡|마약김밥", "korean_restaurant"),

    # ── 24. 카페·디저트·베이커리 ─────────────────────────────
    (r"베이커리|빵집|제과|제빵|베이글", "bakery"),
    (r"카페|커피|coffee|디저트|아이스크림|티하우스|찻집|밀크티", "cafe"),

    # ── 25. 레포츠·온천·스파 ─────────────────────────────────
    (r"패러글라이딩|서핑|스쿠버|스키장|스키리조트|보드|짚라인|루지|"
     r"카약|요트|보트|승마|승마장|스카이다이빙|번지점프|체험$|"
     r"온천$|^.+온천|유성온천|수안보온천|온천랜드|찜질방|스파$|"
     r"녹차탕|해수탕|^.+탕$|율포해수녹차탕"
     , "leisure_sport"),

    # ── 25-2. 숙박·리조트 ────────────────────────────────────
    (r"리조트|호텔|콘도|풀빌라|펜션|모텔|게스트하우스|^워커힐$|^.+호텔$|^.+리조트$",
     "lodging"),

    # ── 25-3. 다리·교량 ──────────────────────────────────────
    (r"다리$|^.+다리$|^.+교$|엑스포다리|영도대교|광안대교|새천년대교", "street_plaza"),

    # ── 26. 차이나타운·외국인 거리 ───────────────────────────
    (r"차이나타운|이태원로|용산기지|"
     r"양림동|예술촌|아트빌리지|아트밸리|아트타운|예술마을|예술의거리|벽화마을|"
     r"창동예술촌", "street_plaza"),

    # ── 27. 빌딩·전망형 건물 ────────────────────────────────
    (r"63빌딩|타워팰리스|롯데월드타워|광화빌딩|^.+빌딩$|광화문빌딩", "tower_observatory"),

    # ── 28. 한옥마을·전통가옥 추가 ──────────────────────────
    (r"^.+촌$|선비촌|민속촌|국악촌|문학촌|"
     r"최참판댁|^.+댁$|^.+종가$|^.+고택$|^.+가옥$", "traditional_village"),
]


def match_sub_category(name: str, hint: str) -> str:
    """이름 + 카테고리힌트로 sub_category 결정."""
    name_clean = (name or "").replace(" ", "")
    for pattern, sub in KEYWORD_RULES:
        if re.search(pattern, name_clean):
            return sub

    # 카테고리힌트로 fallback
    hint = (hint or "").strip()
    if hint == "식당":
        return "general_restaurant"
    if hint == "카페":
        return "cafe"
    if hint == "문화시설":
        return "museum"
    if hint == "쇼핑":
        return "department_store"
    if hint == "레포츠":
        return "leisure_sport"
    if hint == "축제":
        return "default_attraction"

    return "default_attraction"


def confidence(name: str, sub: str) -> str:
    """매칭 확신도 — 후속 검색 우선순위 결정."""
    if sub == "default_attraction":
        return "low"
    # 키워드가 1글자(예: "산", "사")만 매칭된 경우 medium
    name_clean = (name or "").replace(" ", "")
    if len(name_clean) <= 3:
        return "medium"
    return "high"


# ─────────────────────────────────────────────────────────────────
# 1. 322개 → place_hours_322.xlsx 생성
# ─────────────────────────────────────────────────────────────────
def build_place_hours_xlsx() -> tuple[list[dict], Counter]:
    wb_in = load_workbook(INPUT_XLSX, data_only=True)
    ws_in = wb_in["input"]
    headers = [c.value for c in ws_in[1]]

    rows: list[dict] = []
    for r in ws_in.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        rec = dict(zip(headers, r))
        if not (rec.get("여행지명") or "").strip():
            continue
        rows.append(rec)

    sub_counter: Counter[str] = Counter()
    enriched: list[dict] = []
    for rec in rows:
        sub = match_sub_category(rec.get("여행지명") or "", rec.get("카테고리힌트") or "")
        sub_counter[sub] += 1
        meta = SUB_CATEGORIES[sub]
        off_days = meta["off_days"]
        off_str = ",".join(["월","화","수","목","금","토","일"][d] for d in off_days) if off_days else ""

        enriched.append({
            **rec,
            "sub_category":         sub,
            "sub_category_label":   meta["label"],
            "open":                 meta["open"],
            "break_start":          meta["break_start"] or "",
            "break_end":            meta["break_end"] or "",
            "close":                meta["close"],
            "off_days":             off_str,
            "confidence":           confidence(rec.get("여행지명") or "", sub),
            "source":               "category_baseline",
            "note":                 meta["note"],
        })

    # ── 엑셀 작성 ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "place_hours"

    cols = [
        ("plan_id",            10),
        ("day",                 5),
        ("방문순서",             8),
        ("여행지명",            22),
        ("시도",                14),
        ("시군구",              13),
        ("카테고리힌트",         11),
        ("sub_category",        18),
        ("sub_category_label",  22),
        ("open",                 8),
        ("break_start",         11),
        ("break_end",           11),
        ("close",                8),
        ("off_days",            10),
        ("confidence",          11),
        ("source",              18),
        ("note",                40),
    ]
    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
    BODY_FONT   = Font(name="Arial", size=10)
    THIN        = Side(border_style="thin", color="D1D5DB")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    LOW_FILL    = PatternFill("solid", fgColor="FEF3C7")
    MED_FILL    = PatternFill("solid", fgColor="F3F4F6")

    for i, (name, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, rec in enumerate(enriched, 2):
        for ci, (name, _) in enumerate(cols, 1):
            v = rec.get(name)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if name in ("여행지명","sub_category_label","note") else "center")
            if rec["confidence"] == "low":
                cell.fill = LOW_FILL
            elif rec["confidence"] == "medium":
                cell.fill = MED_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT_PLACE)
    print(f"[OK] {OUT_PLACE.relative_to(ROOT)} ({len(enriched)} rows)")
    return enriched, sub_counter


# ─────────────────────────────────────────────────────────────────
# 2. category_hours.xlsx 생성 (카테고리 표준값 테이블)
# ─────────────────────────────────────────────────────────────────
def build_category_std_xlsx(sub_counter: Counter[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "category_hours"

    cols = [
        ("sub_category",   18),
        ("label",          26),
        ("open",            8),
        ("break_start",    11),
        ("break_end",      11),
        ("close",           8),
        ("off_days_idx",   12),
        ("off_days_kr",    12),
        ("count_322",      11),
        ("note",           48),
    ]
    HEADER_FILL = PatternFill("solid", fgColor="065A82")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
    BODY_FONT   = Font(name="Arial", size=10)
    THIN        = Side(border_style="thin", color="D1D5DB")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for i, (n, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=n)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    DAY_KR = ["월","화","수","목","금","토","일"]
    for ri, (sub, meta) in enumerate(SUB_CATEGORIES.items(), 2):
        off_idx = ",".join(str(d) for d in meta["off_days"]) if meta["off_days"] else ""
        off_kr  = ",".join(DAY_KR[d] for d in meta["off_days"]) if meta["off_days"] else ""
        vals = [
            sub, meta["label"], meta["open"], meta["break_start"] or "",
            meta["break_end"] or "", meta["close"],
            off_idx, off_kr, sub_counter.get(sub, 0), meta["note"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if ci in (2,10) else "center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 합계 row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{total_row-1})").font = Font(bold=True, name="Arial")

    wb.save(OUT_STD)
    print(f"[OK] {OUT_STD.relative_to(ROOT)} ({len(SUB_CATEGORIES)} categories)")


# ─────────────────────────────────────────────────────────────────
# 3. hours_db.py 생성 (코드 룩업 모듈)
# ─────────────────────────────────────────────────────────────────
def build_hours_db_py() -> None:
    OUT_PY.parent.mkdir(parents=True, exist_ok=True)

    items_src = []
    for sub, meta in SUB_CATEGORIES.items():
        items_src.append(
            f'    "{sub}": HoursSpec(\n'
            f'        label={meta["label"]!r},\n'
            f'        open_={meta["open"]!r}, close_={meta["close"]!r},\n'
            f'        break_start={meta["break_start"]!r}, break_end={meta["break_end"]!r},\n'
            f'        off_days={meta["off_days"]!r},\n'
            f'    ),'
        )
    rules_src = ",\n".join(f"    ({pat!r}, {sub!r})" for pat, sub in KEYWORD_RULES)

    code = f'''"""hours_db — 카테고리·키워드 기반 운영시간 룩업.

사용:
    from src.data.hours_db import resolve_hours
    spec = resolve_hours("경복궁", "관광지")
    # spec.open_, spec.close_, spec.off_days, spec.break_start, spec.break_end
"""
from __future__ import annotations
import re
from dataclasses import dataclass, field


@dataclass(frozen=True)
class HoursSpec:
    label: str
    open_: str
    close_: str
    break_start: str | None = None
    break_end:   str | None = None
    off_days: list[int] = field(default_factory=list)   # 0=월, ..., 6=일


# ── 카테고리 표준값 (category_hours.xlsx 와 동기화) ──────────
SUB_CATEGORIES: dict[str, HoursSpec] = {{
{chr(10).join(items_src)}
}}

# ── 키워드 → sub_category 매칭 룰 (위에서 아래로 평가) ───────
KEYWORD_RULES: list[tuple[str, str]] = [
{rules_src},
]


def resolve_hours(name: str, hint: str | None = None) -> HoursSpec:
    """장소 이름 + 카테고리힌트 → HoursSpec.

    매칭 우선순위: 키워드 > 카테고리힌트 > default_attraction.
    """
    name_clean = (name or "").replace(" ", "")
    for pattern, sub in KEYWORD_RULES:
        if re.search(pattern, name_clean):
            return SUB_CATEGORIES[sub]

    h = (hint or "").strip()
    fallback = {{
        "식당":     "general_restaurant",
        "카페":     "cafe",
        "문화시설": "museum",
        "쇼핑":     "department_store",
        "레포츠":   "leisure_sport",
        "축제":     "default_attraction",
    }}.get(h, "default_attraction")
    return SUB_CATEGORIES[fallback]


def to_minutes(hhmm: str) -> int:
    h, m = hhmm.split(":")
    return int(h) * 60 + int(m)


def is_open_at(spec: HoursSpec, weekday: int, hhmm: str) -> bool:
    """주어진 요일·시각에 영업 중인지 (주의: 휴무일 + 브레이크 타임 고려)."""
    if weekday in spec.off_days:
        return False
    t = to_minutes(hhmm)
    if not (to_minutes(spec.open_) <= t <= to_minutes(spec.close_)):
        return False
    if spec.break_start and spec.break_end:
        if to_minutes(spec.break_start) <= t < to_minutes(spec.break_end):
            return False
    return True


__all__ = ["HoursSpec", "SUB_CATEGORIES", "KEYWORD_RULES", "resolve_hours", "is_open_at"]
'''
    OUT_PY.write_text(code, encoding="utf-8")
    print(f"[OK] {OUT_PY.relative_to(ROOT)} ({len(SUB_CATEGORIES)} categories, {len(KEYWORD_RULES)} rules)")


# ─────────────────────────────────────────────────────────────────
def main() -> None:
    enriched, counter = build_place_hours_xlsx()
    build_category_std_xlsx(counter)
    build_hours_db_py()

    print()
    print("== sub_category 분포 (322개) ==")
    for sub, n in counter.most_common():
        label = SUB_CATEGORIES[sub]["label"]
        print(f"  {n:4d}  {sub:24s}  {label}")

    low_n = sum(1 for r in enriched if r["confidence"] == "low")
    med_n = sum(1 for r in enriched if r["confidence"] == "medium")
    print()
    print(f"== confidence: low={low_n}, medium={med_n}, high={len(enriched)-low_n-med_n} ==")
    print("→ low/medium 는 후속 검색·수동 보정 우선순위")


if __name__ == "__main__":
    main()
