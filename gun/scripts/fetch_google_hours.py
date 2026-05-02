"""Google Places API (New) — 식당·카페 운영시간 보강.

사용법:
    1. .env 에 GOOGLE_PLACES_API_KEY 추가
    2. python3 gun/scripts/fetch_google_hours.py
       --input  gun/data/recommendations_input.xlsx
       --output gun/data/google_hours.csv

비용 (2025 기준):
    Text Search Pro  : $5 / 1,000 호출
    Place Details Pro: $5 / 1,000 호출
    → 한 POI당 2회 호출 = $10 / 1,000 POI
    → 무료 크레딧 $200/월 = 약 20,000 POI/월
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

# ── 설정 ────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
ENV  = ROOT / ".env"
BASE = "https://places.googleapis.com/v1"

# 식당·카페 카테고리 힌트 (이 프로젝트의 입력 양식 기준)
TARGET_HINTS = {"식당", "카페", "술집", "음식점"}

REQUEST_TIMEOUT = 15.0
SLEEP_BETWEEN   = 0.1   # rate limit 보호

# ── .env 로드 (python-dotenv 없을 때 수동 파싱) ─────────────────
def load_env(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())

load_env(ENV)
API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY")
if not API_KEY:
    print("[error] .env 에 GOOGLE_PLACES_API_KEY 가 없습니다.", file=sys.stderr)
    sys.exit(1)

# ── API 호출 ────────────────────────────────────────────────────
def text_search(query: str, sigungu: str = "") -> str | None:
    """POI 이름으로 검색 → 첫 후보의 place_id 반환."""
    url = f"{BASE}/places:searchText"
    headers = {
        "Content-Type":      "application/json",
        "X-Goog-Api-Key":    API_KEY,
        "X-Goog-FieldMask":  "places.id,places.displayName,places.formattedAddress",
    }
    payload = {
        "textQuery":     f"{query} {sigungu}".strip(),
        "languageCode":  "ko",
        "regionCode":    "KR",
        "pageSize":      1,
    }
    try:
        r = httpx.post(url, headers=headers, json=payload, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        if data.get("places"):
            return data["places"][0]["id"]
    except httpx.HTTPError as e:
        print(f"  [warn] text_search 실패 ({query}): {e}", file=sys.stderr)
    return None


def place_details(place_id: str) -> dict[str, Any] | None:
    """place_id 로 운영시간 조회."""
    url = f"{BASE}/places/{place_id}"
    headers = {
        "X-Goog-Api-Key":    API_KEY,
        "X-Goog-FieldMask":  "id,displayName,regularOpeningHours,currentOpeningHours",
    }
    try:
        r = httpx.get(url, headers=headers,
                      params={"languageCode": "ko"},
                      timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        return r.json()
    except httpx.HTTPError as e:
        print(f"  [warn] place_details 실패 ({place_id}): {e}", file=sys.stderr)
    return None


# ── 운영시간 파싱 ──────────────────────────────────────────────
def parse_periods(periods: list[dict]) -> dict[str, str]:
    """API의 periods 응답을 요일별 'HH:MM-HH:MM' 으로 변환.
    day: 0=일, 1=월, ..., 6=토 (Google 기준)
    """
    DAY_KO = ["일", "월", "화", "수", "목", "금", "토"]
    out: dict[str, str] = {d: "휴무" for d in DAY_KO}
    for p in periods:
        op = p.get("open")
        cl = p.get("close")
        if not op:
            continue
        d = DAY_KO[op["day"]]
        oh, om = op.get("hour", 0), op.get("minute", 0)
        if cl:
            ch, cm = cl.get("hour", 23), cl.get("minute", 59)
            out[d] = f"{oh:02d}:{om:02d}-{ch:02d}:{cm:02d}"
        else:
            out[d] = f"{oh:02d}:{om:02d}-24:00"  # 24시간
    return out


# ── 입력 로드 (Excel) ───────────────────────────────────────────
def load_targets(xlsx_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["input"]
    headers = [c.value for c in ws[1]]
    targets: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        hint = (rec.get("카테고리힌트") or "").strip()
        name = (rec.get("여행지명") or "").strip()
        sgg  = (rec.get("시군구")    or "").strip()
        if hint not in TARGET_HINTS or not name:
            continue
        key = f"{name}|{sgg}"
        if key in seen:
            continue
        seen.add(key)
        targets.append({"name": name, "sigungu": sgg, "category": hint})
    return targets


# ── 메인 ────────────────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default=str(ROOT / "gun/data/recommendations_input.xlsx"))
    ap.add_argument("--output", default=str(ROOT / "gun/data/google_hours.csv"))
    ap.add_argument("--limit",  type=int, default=None,
                    help="처음 N개만 처리 (테스트용)")
    args = ap.parse_args()

    targets = load_targets(Path(args.input))
    if args.limit:
        targets = targets[:args.limit]

    print(f"[start] 식당·카페 {len(targets)}개 운영시간 조회 시작")
    print(f"        예상 비용: ${len(targets)*2*5/1000:.2f} (호출 {len(targets)*2}회)")

    fieldnames = [
        "name", "sigungu", "category",
        "place_id", "google_name",
        "월", "화", "수", "목", "금", "토", "일",
        "weekday_descriptions",
    ]
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()

        ok, fail = 0, 0
        for i, tgt in enumerate(targets, 1):
            print(f"[{i}/{len(targets)}] {tgt['name']} ({tgt['sigungu']})", end=" ... ")
            pid = text_search(tgt["name"], tgt["sigungu"])
            time.sleep(SLEEP_BETWEEN)
            if not pid:
                print("not_found")
                w.writerow({**tgt, "place_id": "", "google_name": "",
                            "weekday_descriptions": "NOT_FOUND"})
                fail += 1
                continue

            det = place_details(pid)
            time.sleep(SLEEP_BETWEEN)
            if not det:
                print("details_fail")
                w.writerow({**tgt, "place_id": pid, "google_name": "",
                            "weekday_descriptions": "DETAILS_FAIL"})
                fail += 1
                continue

            hours = det.get("regularOpeningHours", {}) or {}
            periods = hours.get("periods", [])
            descs   = hours.get("weekdayDescriptions", [])
            day_map = parse_periods(periods) if periods else {}

            row = {
                **tgt,
                "place_id":    pid,
                "google_name": det.get("displayName", {}).get("text", ""),
                "월": day_map.get("월", ""),
                "화": day_map.get("화", ""),
                "수": day_map.get("수", ""),
                "목": day_map.get("목", ""),
                "금": day_map.get("금", ""),
                "토": day_map.get("토", ""),
                "일": day_map.get("일", ""),
                "weekday_descriptions": " | ".join(descs),
            }
            w.writerow(row)
            ok += 1
            print(f"OK · {hours.get('weekdayDescriptions', ['?'])[0] if descs else '시간 미공개'}")

    print()
    print(f"[done] 성공 {ok}, 실패 {fail}")
    print(f"       결과: {args.output}")


if __name__ == "__main__":
    main()
