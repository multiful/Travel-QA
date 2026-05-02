"""recommendations_input.xlsx 입력 양식 + 예시 생성.

실행: python3 gun/scripts/create_input_template.py
산출: gun/data/recommendations_input.xlsx (3 시트)
       - input    : 빈 양식 + 예시 5 row + 데이터 검증(Validation)
       - guide    : 컬럼 의미 안내
       - 예시일정 : 완성된 일정 1개 샘플
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

GUN_DATA = Path(__file__).resolve().parents[1] / "data"
GUN_DATA.mkdir(parents=True, exist_ok=True)
OUTPUT = GUN_DATA / "recommendations_input.xlsx"

# ── 컬럼 정의 ─────────────────────────────────────────────────────────
INPUT_COLS = [
    ("source",      "추천 출처 (트리플/마이리얼트립/블로그명)"),
    ("plan_id",     "일정 식별자 (T001, MRT042 등 자유)"),
    ("시도",         "전라남도, 서울특별시 등"),
    ("시군구",       "강남구, 경주시, 해운대구 등"),
    ("여행기간",     "당일치기 / 1박2일 / 2박3일"),
    ("일자",         "YYYY-MM-DD (이 day 의 날짜)"),
    ("day",          "1, 2, 3 (다일 일정의 day 번호)"),
    ("방문순서",     "1, 2, 3, 4, ..."),
    ("여행지명",     "정확한 한국어 이름 (오타 주의 — 매칭률 핵심)"),
    ("카테고리힌트", "(선택) 카페/식당/관광지/술집 등 — 매칭 보조용"),
]

# ── 예시 row (1개 일정 = 6 places) ───────────────────────────────────
EXAMPLE_ROWS = [
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 1, "경복궁",       "관광지"),
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 2, "광화문",       "관광지"),
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 3, "토속촌삼계탕", "식당"),
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 4, "북촌한옥마을", "관광지"),
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 5, "오니오니",      "카페"),
    ("트리플",        "T001", "서울특별시", "종로구", "당일치기", "2025-05-01", 1, 6, "광장시장",     "식당"),
]

# ── 시도 드롭다운 옵션 ───────────────────────────────────────────────
SIDO_OPTIONS = [
    "서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시",
    "울산광역시", "세종특별자치시", "경기도", "강원특별자치도", "충청북도", "충청남도",
    "전북특별자치도", "전라남도", "경상북도", "경상남도", "제주특별자치도",
]

PERIOD_OPTIONS = ["당일치기", "1박2일", "2박3일", "3박4일"]
CATEGORY_HINTS  = ["관광지", "문화시설", "식당", "카페", "술집", "쇼핑", "레포츠", "축제", "숙박"]


def make_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── 시트 1: input ──
    ws = wb.active
    ws.title = "input"

    # 헤더
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (name, _) in enumerate(INPUT_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 예시 5 row
    example_fill = PatternFill("solid", fgColor="E7E6E6")
    for row_idx, row in enumerate(EXAMPLE_ROWS, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    # 컬럼 폭 자동 조절
    widths = [12, 10, 14, 12, 12, 12, 6, 8, 22, 12]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 데이터 검증 (드롭다운)
    # 시도 (C열)
    dv_sido = DataValidation(type="list",
                             formula1=f'"{",".join(SIDO_OPTIONS)}"',
                             allow_blank=True)
    dv_sido.add(f"C2:C1000")
    ws.add_data_validation(dv_sido)

    # 여행기간 (E열)
    dv_period = DataValidation(type="list",
                               formula1=f'"{",".join(PERIOD_OPTIONS)}"',
                               allow_blank=True)
    dv_period.add(f"E2:E1000")
    ws.add_data_validation(dv_period)

    # 카테고리힌트 (J열)
    dv_cat = DataValidation(type="list",
                            formula1=f'"{",".join(CATEGORY_HINTS)}"',
                            allow_blank=True)
    dv_cat.add(f"J2:J1000")
    ws.add_data_validation(dv_cat)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # ── 시트 2: guide ──
    ws_guide = wb.create_sheet("guide")
    ws_guide.cell(row=1, column=1, value="컬럼명").font = Font(bold=True, size=12)
    ws_guide.cell(row=1, column=2, value="설명").font = Font(bold=True, size=12)
    for r, (name, desc) in enumerate(INPUT_COLS, 2):
        ws_guide.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws_guide.cell(row=r, column=2, value=desc)

    # 추가 가이드 박스
    ws_guide.cell(row=len(INPUT_COLS) + 4, column=1,
                  value="📋 입력 시 주의사항").font = Font(bold=True, size=12)
    notes = [
        "1. 같은 일정의 모든 장소는 plan_id가 같아야 합니다 (예: T001 5개 row).",
        "2. 여행지명은 공식 표기 그대로 — 매칭 정확도가 핵심 (오타 시 not_found 처리).",
        "3. 다일 일정은 day=1, 2, 3 으로 분리 (각 day마다 별도 row 그룹).",
        "4. 카테고리힌트는 선택. 비워둬도 자동 매칭됨.",
        "5. 일정당 권장 장소 수: 5~6개 (식당2 + 카페1 + 관광지2~3).",
        "6. 매칭 실패율 10% 초과하면 build_itinerary_excel.py가 경고 출력.",
        "",
        "🔧 작업 흐름",
        "  ① 이 input 시트에 일정 채우기 (저장)",
        "  ② 터미널에서: python3 gun/scripts/build_itinerary_excel.py",
        "  ③ gun/data/itinerary_results_YYYYMMDD.xlsx 결과 확인",
        "  ④ match_failed.csv 가 생성되면 이름 보정 후 ②~③ 재실행",
    ]
    for i, note in enumerate(notes, 1):
        ws_guide.cell(row=len(INPUT_COLS) + 4 + i, column=1, value=note)

    ws_guide.column_dimensions["A"].width = 16
    ws_guide.column_dimensions["B"].width = 60

    # ── 시트 3: 예시일정 ──
    ws_ex = wb.create_sheet("예시일정_T001")
    for col_idx, (name, _) in enumerate(INPUT_COLS, 1):
        cell = ws_ex.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(EXAMPLE_ROWS, 2):
        for col_idx, value in enumerate(row, 1):
            ws_ex.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, w in enumerate(widths, 1):
        ws_ex.column_dimensions[get_column_letter(col_idx)].width = w
    ws_ex.freeze_panes = "A2"

    return wb


def main():
    wb = make_workbook()
    wb.save(OUTPUT)
    print(f"[OK] 입력 양식 생성: {OUTPUT}")
    print(f"     시트: input(예시 5row), guide, 예시일정_T001")
    print(f"     컬럼: {len(INPUT_COLS)}개")
    print(f"\n다음:")
    print(f"  1. {OUTPUT} 열어서 'input' 시트에 일정 채우기")
    print(f"  2. python3 gun/scripts/build_itinerary_excel.py 실행")


if __name__ == "__main__":
    main()
